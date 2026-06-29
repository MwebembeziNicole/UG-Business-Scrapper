"""
Excel exporter — one sheet per platform, saved to the exports/ folder.
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

EXPORTS_DIR = config.EXPORTS_DIR
os.makedirs(EXPORTS_DIR, exist_ok=True)

PLATFORM_COLORS = {
    "jiji":        "FF6B35",
    "instagram":   "C13584",
    "yellowpages": "F5A623",
    "twitter":     "1DA1F2",
    "tiktok":      "010101",
}

COLUMNS = ["business_name", "phone", "email", "category", "location", "website", "facebook", "source_url"]
HEADERS = ["Business Name", "Phone", "Email", "Category", "Location", "Website", "Facebook", "URL"]


def export_platform(platform: str, records: list) -> str:
    if not records:
        df = pd.DataFrame(columns=COLUMNS)
    else:
        df = pd.DataFrame(records)
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = ""
        df = df[COLUMNS]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"{platform}_businesses_{timestamp}.xlsx"
    filepath  = os.path.join(EXPORTS_DIR, filename)

    df.to_excel(filepath, index=False, sheet_name=platform.title())

    wb = load_workbook(filepath)
    ws = wb.active

    hex_color    = PLATFORM_COLORS.get(platform, "4F81BD")
    header_fill  = PatternFill("solid", fgColor=hex_color)
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    alt_fill     = PatternFill("solid", fgColor="F2F2F2")
    border_color = "D0D0D0"
    thin_border  = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    # Style header row
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
        cell.value     = HEADERS[cell.column - 1]

    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for cell in row:
            if fill:
                cell.fill = fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Auto-fit column widths
    col_widths = [34, 16, 28, 20, 24, 34, 34, 50]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    # Insert title row
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value     = f"{platform.upper()} Uganda Business Contacts — {datetime.now().strftime('%B %d, %Y')}"
    title_cell.font      = Font(bold=True, size=13, color=hex_color)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    wb.save(filepath)
    return filepath


def export_all(records_by_platform: dict) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"all_platforms_{timestamp}.xlsx"
    filepath  = os.path.join(EXPORTS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for platform, records in records_by_platform.items():
            df = pd.DataFrame(records) if records else pd.DataFrame(columns=COLUMNS)
            for col in COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            df = df[COLUMNS]
            df.columns = HEADERS
            df.to_excel(writer, sheet_name=platform.title(), index=False)

    wb = load_workbook(filepath)
    for platform in records_by_platform:
        if platform.title() not in wb.sheetnames:
            continue
        ws = wb[platform.title()]
        hex_color   = PLATFORM_COLORS.get(platform, "4F81BD")
        header_fill = PatternFill("solid", fgColor=hex_color)
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_widths = [34, 16, 28, 20, 24, 34, 34, 50]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath


# ── Dated daily snapshot ──────────────────────────────────────────────────────
# Guarantees one clean, predictable "today" workbook per day, regardless of how
# many businesses are brand-new. Re-running on the same day overwrites that day's
# file, so there is always exactly one current daily list to export.

DAILY_DIR = config.DAILY_EXPORTS_DIR


def _today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _style_sheet(ws, hex_color: str = "0D47A1"):
    header_fill = PatternFill("solid", fgColor=hex_color)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, width in enumerate([34, 16, 28, 20, 24, 34, 34, 50], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _frame(records: list):
    df = pd.DataFrame(records) if records else pd.DataFrame(columns=COLUMNS)
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMNS]
    df.columns = HEADERS
    return df


def export_daily(records_by_platform: dict, new_by_platform: dict = None) -> str:
    """
    Write one dated workbook for today:
      • a "New Today" sheet (businesses collected today, across all platforms)
      • one sheet per platform with the full current dataset

    Saved to exports/daily/businesses_<YYYY-MM-DD>.xlsx and overwritten on each
    run that day, so the investigator always has a single up-to-date daily list.
    """
    os.makedirs(DAILY_DIR, exist_ok=True)
    date     = _today_str()
    filepath = os.path.join(DAILY_DIR, f"businesses_{date}.xlsx")

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        if new_by_platform is not None:
            rows = [r for recs in new_by_platform.values() for r in recs]
            _frame(rows).to_excel(writer, sheet_name="New Today", index=False)
        for platform, records in records_by_platform.items():
            _frame(records).to_excel(writer, sheet_name=platform.title()[:31], index=False)

    wb = load_workbook(filepath)
    for ws in wb.worksheets:
        _style_sheet(ws)
    wb.save(filepath)
    return filepath


def get_daily_export(date: str = None) -> str | None:
    date = date or _today_str()
    fp = os.path.join(DAILY_DIR, f"businesses_{date}.xlsx")
    return fp if os.path.exists(fp) else None


def get_latest_export(platform: str) -> str | None:
    files = [
        os.path.join(EXPORTS_DIR, f)
        for f in os.listdir(EXPORTS_DIR)
        if f.startswith(f"{platform}_businesses_") and f.endswith(".xlsx")
    ]
    if not files:
        return None
    return max(files, key=os.path.getmtime)
